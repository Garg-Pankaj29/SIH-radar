#!/usr/bin/env python3
"""
SIH 2026 Problem Statement Traffic Tracker
--------------------------------------------
Fetches the live SIH 2026 problem-statement dataset (idea-submission counts
per PS, refreshed daily from sih.gov.in) and builds a ranked Excel report
showing which PS are filling up fast (high competition) vs still wide open
(low competition), so you can pick strategically.

Data source: https://github.com/vedantchalke36/sih-2026-problem-statements
(community-maintained daily scrape of the official sih.gov.in portal).

Usage:
    python sih_traffic_tracker.py [output.xlsx]

Re-run this anytime (e.g. weekly) as the deadline approaches -- the
"ideas submitted" counters only become meaningful once colleges start
uploading team ideas nationally.
"""
import json
import re
import sys
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA_URL = "https://raw.githubusercontent.com/vedantchalke36/sih-2026-problem-statements/main/data/sih2026_ps.json"

def fetch_data():
    with urllib.request.urlopen(DATA_URL, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))

def parse_ideas(ideas_str):
    """'37/500' -> (37, 500). Falls back to (0, 500) if unparseable."""
    m = re.match(r"(\d+)\s*/\s*(\d+)", str(ideas_str or ""))
    if m:
        return int(m.group(1)), int(m.group(2))
    return 0, 500

def traffic_label(pct):
    if pct >= 50:
        return "High"
    elif pct >= 10:
        return "Medium"
    elif pct > 0:
        return "Low"
    else:
        return "None yet"

def build_workbook(records, out_path):
    rows = []
    for d in records:
        submitted, cap = parse_ideas(d.get("ideas"))
        pct = round(100 * submitted / cap, 1) if cap else 0.0
        rows.append({
            "ps_number": d.get("ps_number", ""),
            "title": d.get("title", ""),
            "org": d.get("org", ""),
            "category": d.get("category", ""),
            "theme": d.get("theme", ""),
            "deadline": d.get("deadline", ""),
            "submitted": submitted,
            "cap": cap,
            "pct": pct,
            "traffic": traffic_label(pct),
            "link": f"https://sih.gov.in/sih2026PS" ,
        })

    # sort high -> low competition
    rows.sort(key=lambda r: (-r["pct"], r["ps_number"]))

    wb = openpyxl.Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["PS Number", "Title", "Organization", "Category", "Theme",
               "Deadline", "Ideas Submitted", "Cap", "Fill %", "Competition Level"]
    widths = [12, 45, 30, 11, 24, 16, 14, 8, 9, 16]

    def write_sheet(ws, data_rows, title):
        ws.title = title
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        for ri, r in enumerate(data_rows, start=2):
            vals = [r["ps_number"], r["title"], r["org"], r["category"], r["theme"],
                    r["deadline"], r["submitted"], r["cap"], r["pct"]/100, r["traffic"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = normal_font
                c.border = border
                c.alignment = left if ci in (2, 3, 5) else center
                if ci == 9:
                    c.number_format = "0.0%"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if data_rows:
            last = len(data_rows) + 1
            rng = f"I2:I{last}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=PatternFill("solid", fgColor="F8696B")))
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.1","0.4999"], fill=PatternFill("solid", fgColor="FFEB84")))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.1"], fill=PatternFill("solid", fgColor="63BE7B")))
        ws.auto_filter.ref = f"A1:J{len(data_rows)+1}"

    ws1 = wb.active
    write_sheet(ws1, rows, "All PS Ranked")

    ws2 = wb.create_sheet()
    write_sheet(ws2, [r for r in rows if r["pct"] > 0][:30] or rows[:30], "High Traffic (Top 30)")

    ws3 = wb.create_sheet()
    low_rows = sorted(rows, key=lambda r: (r["pct"], r["ps_number"]))[:30]
    write_sheet(ws3, low_rows, "Low Traffic (Best Odds)")

    # Theme summary sheet
    ws4 = wb.create_sheet("By Theme")
    theme_counts = {}
    for r in rows:
        t = theme_counts.setdefault(r["theme"], {"count": 0, "submitted": 0})
        t["count"] += 1
        t["submitted"] += r["submitted"]
    ws4.append(["Theme", "Number of PS", "Total Ideas Submitted", "Avg Ideas / PS"])
    for i, h in enumerate(["Theme", "Number of PS", "Total Ideas Submitted", "Avg Ideas / PS"], 1):
        c = ws4.cell(row=1, column=i)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
    for ri, (theme, v) in enumerate(sorted(theme_counts.items(), key=lambda x: -x[1]["count"]), start=2):
        avg = round(v["submitted"] / v["count"], 2) if v["count"] else 0
        vals = [theme, v["count"], v["submitted"], avg]
        for ci, val in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.font = normal_font; c.border = border
            c.alignment = left if ci == 1 else center
    ws4.column_dimensions["A"].width = 38
    for col in ["B","C","D"]:
        ws4.column_dimensions[col].width = 18
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:D{len(theme_counts)+1}"

    # Notes sheet
    ws5 = wb.create_sheet("Read Me")
    notes = [
        ("SIH 2026 Problem Statement Traffic Tracker", ""),
        (f"Generated on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Data source", "sih.gov.in (official), via community daily scrape on GitHub"),
        ("", ""),
        ("What 'traffic' means here", "SIH does not publish page-view/click analytics. This report uses the official "
         "'ideas submitted' counter per problem statement (capped at 500 nationally) as the real "
         "popularity/competition signal -- it reflects committed teams, not casual clicks."),
        ("Competition Level", "High (>=50% filled) = crowded, harder to stand out. "
         "Medium (10-49%) = moderate competition. Low/None (<10%) = best odds, less competition."),
        ("Important", "Counts only become meaningful as national idea-submission ramps up "
         "(through Sept 2026). Early in the cycle most PS will show 0 -- re-run this script "
         "closer to your PS deadline for a real read."),
        ("How to refresh", "Run: python sih_traffic_tracker.py  -- it re-fetches the latest data "
         "and rebuilds this file."),
    ]
    for ri, (a, b) in enumerate(notes, start=1):
        ws5.cell(row=ri, column=1, value=a).font = Font(name="Arial", bold=(ri==1), size=13 if ri==1 else 10)
        c2 = ws5.cell(row=ri, column=2, value=b)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws5.column_dimensions["A"].width = 26
    ws5.column_dimensions["B"].width = 90

    wb.move_sheet("Read Me", offset=-4)  # put Read Me first
    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    out_path = sys.argv[1] if len(sys.argv) > 1 else "SIH2026_PS_Traffic_Report.xlsx"
    print("Fetching latest SIH 2026 problem statement data...")
    records = fetch_data()
    print(f"Loaded {len(records)} problem statements.")
    build_workbook(records, out_path)
    print(f"Report saved to {out_path}")
